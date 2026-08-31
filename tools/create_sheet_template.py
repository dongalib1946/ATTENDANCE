from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "01a01d7b-6591-7ff0-ab17-13f739cca685"
OUTPUT_PATH = OUTPUT_DIR / "근로학생_출퇴근_상호대차_구글시트_템플릿.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="0F766E")
SUBTLE_FILL = PatternFill("solid", fgColor="EEF6F2")
NOTE_FILL = PatternFill("solid", fgColor="F7F9F8")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=15, bold=True, color="10231F")
BODY_FONT = Font(size=10, color="10231F")
MUTED_FONT = Font(size=9, color="5F6D68")
THIN_BORDER = Border(bottom=Side(style="thin", color="D8E0DC"))


def style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_body(ws, min_row, max_row, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center")


def add_table(ws, ref, name):
    table = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def build_workbook():
    wb = Workbook()
    attendance = wb.active
    attendance.title = "출퇴근기록"
    roster = wb.create_sheet("학생명부")
    schedule = wb.create_sheet("시간표")
    settings = wb.create_sheet("설정")

    attendance.append(["이름", "시간", "구분", "층"])
    attendance.append(["홍길동", "2026-08-20 09:00:00", "출근", "4층"])
    attendance.append(["홍길동", "2026-08-20 17:00:00", "퇴근", "4층"])
    attendance.append(["김학생", "2026-08-20 09:05:00", "출근", "5층"])
    style_header(attendance[1])
    style_body(attendance, 2, 200, 4)
    attendance.freeze_panes = "A2"
    attendance.column_dimensions["A"].width = 16
    attendance.column_dimensions["B"].width = 22
    attendance.column_dimensions["C"].width = 12
    attendance.column_dimensions["D"].width = 10
    add_table(attendance, "A1:D200", "AttendanceLog")

    roster.append(["층", "이름", "사용여부", "비고"])
    roster_rows = [
        ["4층", "홍길동", "Y", "예시 행입니다. 실제 학생명으로 바꿔 주세요."],
        ["4층", "이근로", "Y", ""],
        ["5층", "김학생", "Y", ""],
        ["5층", "박근로", "Y", ""],
        ["5층", "비활성예시", "N", "N이면 앱 이름 목록에서 제외됩니다."],
    ]
    for row in roster_rows:
        roster.append(row)
    style_header(roster[1])
    style_body(roster, 2, 205, 4)
    roster.freeze_panes = "A2"
    roster.column_dimensions["A"].width = 10
    roster.column_dimensions["B"].width = 16
    roster.column_dimensions["C"].width = 12
    roster.column_dimensions["D"].width = 42
    add_table(roster, "A1:D205", "StudentRoster")

    floor_validation = DataValidation(type="list", formula1='"4층,5층"', allow_blank=False)
    active_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    roster.add_data_validation(floor_validation)
    roster.add_data_validation(active_validation)
    floor_validation.add("A2:A205")
    active_validation.add("C2:C205")

    schedule.append(["요일", "1타임", "2타임", "3타임", "비고", "사용여부"])
    schedule_rows = [
        ["월", "", "", "", "", "Y"],
        ["화", "", "", "", "", "Y"],
        ["수", "", "", "", "", "Y"],
        ["목", "", "", "", "", "Y"],
        ["금", "", "", "", "", "Y"],
    ]
    for row in schedule_rows:
        schedule.append(row)

    style_header(schedule[1])
    style_body(schedule, 2, 6, 6)
    schedule.freeze_panes = "A2"
    schedule.column_dimensions["A"].width = 10
    schedule.column_dimensions["B"].width = 18
    schedule.column_dimensions["C"].width = 18
    schedule.column_dimensions["D"].width = 18
    schedule.column_dimensions["E"].width = 42
    schedule.column_dimensions["F"].width = 12
    add_table(schedule, "A1:F6", "InterloanSchedule")

    weekday_validation = DataValidation(type="list", formula1='"월,화,수,목,금"', allow_blank=False)
    schedule_active_validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    schedule.add_data_validation(weekday_validation)
    schedule.add_data_validation(schedule_active_validation)
    weekday_validation.add("A2:A6")
    schedule_active_validation.add("F2:F6")

    settings["A1"] = "근로학생 출퇴근 QR 앱 설정 메모"
    settings["A1"].font = TITLE_FONT
    settings["A3"] = "시트 구조"
    settings["A4"] = "1번째 시트 출퇴근기록: 이름 / 시간 / 구분 / 층"
    settings["A5"] = "2번째 시트 학생명부: 층 / 이름 / 사용여부 / 비고"
    settings["A6"] = "3번째 시트 시간표: 요일 / 1타임 / 2타임 / 3타임 / 비고 / 사용여부"
    settings["A7"] = "학생명부 사용 방법"
    settings["A8"] = "층은 4층 또는 5층을 선택합니다."
    settings["A9"] = "사용여부가 Y인 학생만 앱 이름 목록에 표시됩니다."
    settings["A10"] = "사용여부를 N으로 바꾸면 과거 기록은 유지하고 선택 목록에서만 제외됩니다."
    settings["A12"] = "시간표 사용 방법"
    settings["A13"] = "월~금 각 요일 행에 상호대차 1타임, 2타임, 3타임 담당 학생명을 입력합니다."
    settings["A14"] = "디스플레이 화면에는 오늘 요일의 3타임 담당자만 표시됩니다."
    settings["A15"] = "사용여부가 N인 요일 행은 디스플레이에서 비활성 안내로 표시됩니다."
    settings["A17"] = "Apps Script 배포 후"
    settings["A18"] = "Code.gs의 SPREADSHEET_ID에는 이 구글시트의 ID를 입력합니다."
    settings["A19"] = "Netlify 환경변수에 GOOGLE_APPS_SCRIPT_URL, ATTENDANCE_PROXY_SECRET, ATTENDANCE_QR_SECRET을 저장합니다."
    settings.merge_cells("A1:D1")
    settings["A3"].fill = SUBTLE_FILL
    settings["A7"].fill = SUBTLE_FILL
    settings["A12"].fill = SUBTLE_FILL
    settings["A17"].fill = SUBTLE_FILL
    for row in range(3, 20):
      settings[f"A{row}"].font = BODY_FONT if row not in (3, 7, 12, 17) else Font(bold=True, color="10231F")
      settings[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")
    settings.column_dimensions["A"].width = 86
    settings.column_dimensions["B"].width = 12
    settings.sheet_view.showGridLines = False
    settings["A21"] = "주의"
    settings["A22"] = "민감한 값은 브라우저 파일에 넣지 말고 Netlify 환경변수와 Apps Script 스크립트 속성에 저장합니다."
    settings["A21"].fill = NOTE_FILL
    settings["A21"].font = Font(bold=True, color="10231F")
    settings["A22"].font = MUTED_FONT
    settings["A22"].alignment = Alignment(wrap_text=True, vertical="center")

    for ws in (attendance, roster, schedule, settings):
        ws.sheet_view.showGridLines = False

    return wb


def verify(path):
    wb = load_workbook(path)
    assert wb.sheetnames == ["출퇴근기록", "학생명부", "시간표", "설정"]
    assert [cell.value for cell in wb["출퇴근기록"][1][:4]] == ["이름", "시간", "구분", "층"]
    assert [cell.value for cell in wb["학생명부"][1][:4]] == ["층", "이름", "사용여부", "비고"]
    assert [cell.value for cell in wb["시간표"][1][:6]] == ["요일", "1타임", "2타임", "3타임", "비고", "사용여부"]
    assert wb["학생명부"]["A2"].value == "4층"
    assert wb["학생명부"]["C6"].value == "N"
    assert wb["시간표"]["A2"].value == "월"
    assert wb["시간표"]["A6"].value == "금"
    assert wb["시간표"]["B2"].value is None
    assert len(wb["학생명부"].data_validations.dataValidation) == 2
    assert len(wb["시간표"].data_validations.dataValidation) == 2


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    verify(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
